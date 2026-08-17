"""ETF 4주 추이 — DB 이력 조회 · processed 보강 · 렌더(markdown/Excel).

추이는 **기존 수집분 재사용**이다. ETF 수급은 이미 weekly_stock(market='ETF')에 쌓이므로
KRX를 다시 부르지 않는다.
"""
import pandas as pd
import pytest
from openpyxl import Workbook

from modules import etf_section, exporter
from modules.database import Database


@pytest.fixture
def db_with_etf_history(tmp_path):
    """3주치 ETF 이력. 오래된 주 → 최신 주 순서가 추이의 계약이다."""
    db = Database(str(tmp_path / "history.db"))
    weeks = {
        "2026-07-31": [("069500", 100.0, 10.0), ("229200", -50.0, -5.0)],
        "2026-08-07": [("069500", 200.0, 20.0), ("229200", -60.0, -6.0)],
        "2026-08-14": [("069500", 300.0, 30.0), ("229200", -70.0, -7.0)],
    }
    for week_date, rows in weeks.items():
        etf = pd.DataFrame(
            [{"티커": t, "1주기관매매": inst, "1주외국인매매": fore} for t, inst, fore in rows]
        )
        db.upsert_week(
            week_date=week_date, market_data={}, cap_weight_data=pd.DataFrame(),
            sector_data=pd.DataFrame(), max_weeks=52,
            processed={"etf_flows": etf, "flow_source": "krx"},
        )
    return db


def test_ETF_이력은_한_번의_조회로_전_종목을_돌려준다(db_with_etf_history):
    """ETF는 100종목대라 종목별로 돌면 쿼리가 그만큼 늘어난다."""
    history = db_with_etf_history.get_etf_flow_history(n_weeks=4)

    assert set(history["ticker"]) == {"069500", "229200"}
    kodex200 = history[history["ticker"] == "069500"].sort_values("week_date")
    assert list(kodex200["inst_net_1w"]) == [100.0, 200.0, 300.0]
    assert list(kodex200["week_date"]) == ["2026-07-31", "2026-08-07", "2026-08-14"]


def test_요청한_주_수만큼만_돌려준다(db_with_etf_history):
    history = db_with_etf_history.get_etf_flow_history(n_weeks=2)

    assert sorted(set(history["week_date"])) == ["2026-08-07", "2026-08-14"]


def test_이력이_없으면_빈_결과다(tmp_path):
    db = Database(str(tmp_path / "empty.db"))

    assert db.get_etf_flow_history(n_weeks=4).empty


def test_보강은_ETF별_추이를_오래된_주부터_붙인다(db_with_etf_history):
    """main._enrich_from_db와 같은 자리 — 렌더러는 읽기만 하게 한다."""
    from main import _enrich_etf_trend

    flows = pd.DataFrame({
        "티커": ["069500", "229200"],
        "종목명": ["KODEX 200", "KODEX 코스닥150"],
        "1주기관매매": [300.0, -70.0],
        "1주외국인매매": [30.0, -7.0],
    })

    enriched = _enrich_etf_trend(db_with_etf_history, flows, n_weeks=4)

    row = enriched[enriched["티커"] == "069500"].iloc[0]
    assert row["기관4주추이"] == [100.0, 200.0, 300.0]
    assert row["외국인4주추이"] == [10.0, 20.0, 30.0]


def test_이력에_없는_ETF는_빈_추이를_받는다(db_with_etf_history):
    from main import _enrich_etf_trend

    flows = pd.DataFrame({
        "티커": ["999999"], "종목명": ["신규 ETF"],
        "1주기관매매": [5.0], "1주외국인매매": [1.0],
    })

    enriched = _enrich_etf_trend(db_with_etf_history, flows, n_weeks=4)

    assert enriched.iloc[0]["기관4주추이"] == []


def test_섹션은_추이_컬럼이_있으면_표에_싣는다():
    flows = pd.DataFrame({
        "티커": ["069500"], "종목명": ["KODEX 200"],
        "1주외국인매매": [30.0], "1주기관매매": [300.0],
        "기관4주추이": [[100.0, 200.0, 300.0]],
        "외국인4주추이": [[10.0, 20.0, 30.0]],
    })

    md = etf_section.build_etf_section(flows, None)

    assert "4주 추이" in md
    assert "+100.0 → +200.0 → +300.0" in md


def test_섹션은_추이_컬럼이_없으면_기존_표를_그대로_낸다():
    """DB 이력이 아직 없는 첫 주에도 섹션이 깨지면 안 된다."""
    flows = pd.DataFrame({
        "티커": ["069500"], "종목명": ["KODEX 200"],
        "1주외국인매매": [30.0], "1주기관매매": [300.0],
    })

    md = etf_section.build_etf_section(flows, None)

    assert md.startswith("## ETF 수급")
    assert "4주 추이" not in md


def test_Excel에_ETF_시트가_생긴다(tmp_path):
    processed = {
        "etf_flows": pd.DataFrame({
            "티커": ["069500"], "종목명": ["KODEX 200"],
            "1주기관매매": [300.0], "1주외국인매매": [30.0],
            "기관4주추이": [[100.0, 200.0, 300.0]],
            "외국인4주추이": [[10.0, 20.0, 30.0]],
        }),
    }
    workbook = Workbook()
    sheet = workbook.active

    exporter._write_etf_sheet(sheet, processed.get("etf_flows"), "2026-08-14")

    values = [[cell.value for cell in row] for row in sheet.iter_rows()]
    flat = [str(v) for row in values for v in row if v is not None]
    assert any("ETF" in v for v in flat)
    assert "KODEX 200" in flat
    assert "+100.0 → +200.0 → +300.0" in flat


def test_ETF가_없으면_시트를_비워두되_깨지지_않는다():
    workbook = Workbook()
    sheet = workbook.active

    exporter._write_etf_sheet(sheet, pd.DataFrame(), "2026-08-14")

    flat = [str(c.value) for row in sheet.iter_rows() for c in row if c.value is not None]
    assert any("데이터 없음" in v for v in flat)
