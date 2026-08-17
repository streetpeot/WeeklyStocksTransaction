"""
exporter.py - Excel 멀티탭 저장 모듈
5탭 구성: 코스피 / 코스닥 / 증시정보 / 코스피시총비중 / 코스닥시총비중
"""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    numbers,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────
# 스타일 상수
# ─────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEADER_FONT = Font(bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
NUM_FORMAT_COMMA = "#,##0"
NUM_FORMAT_COMMA2 = "#,##0.00"
NUM_FORMAT_PCT = "0.00%"
NUM_FORMAT_PCT2 = "0.00"


# ─────────────────────────────────────────
# 컬럼 정의
# ─────────────────────────────────────────

STOCK_COLUMNS = [
    ("순위",            "순위",              10,  NUM_FORMAT_COMMA),
    ("티커",            "티커",              8,   "@"),
    ("종목명",           "종목명",             18,  "@"),
    ("시가총액(억)",     "시가총액",           14,  NUM_FORMAT_COMMA),
    ("거래대금(억)",     "거래대금(억)",        14,  NUM_FORMAT_COMMA),
    ("PER",             "PER",               8,   NUM_FORMAT_COMMA2),
    ("PBR",             "PBR",               8,   NUM_FORMAT_COMMA2),
    ("배당수익률(%)",    "배당수익률",          12,  NUM_FORMAT_COMMA2),
    ("1주등락률(%)",     "1주등락률",           12,  NUM_FORMAT_COMMA2),
    ("1개월등락률(%)",   "1개월등락률",         14,  NUM_FORMAT_COMMA2),
    ("3개월등락률(%)",   "3개월등락률",         14,  NUM_FORMAT_COMMA2),
    ("6개월등락률(%)",   "6개월등락률",         14,  NUM_FORMAT_COMMA2),
    ("1주기관매매(억)",  "1주기관매매",         16,  NUM_FORMAT_COMMA),
    ("2주기관매매(억)",  "2주기관매매",         16,  NUM_FORMAT_COMMA),
    ("1개월기관매매(억)", "1개월기관매매",       18,  NUM_FORMAT_COMMA),
    ("3개월기관매매(억)", "3개월기관매매",       18,  NUM_FORMAT_COMMA),
    ("1주외국인매매(억)", "1주외국인매매",       18,  NUM_FORMAT_COMMA),
    ("2주외국인매매(억)", "2주외국인매매",       18,  NUM_FORMAT_COMMA),
    ("1개월외국인매매(억)", "1개월외국인매매",   20,  NUM_FORMAT_COMMA),
    ("3개월외국인매매(억)", "3개월외국인매매",   20,  NUM_FORMAT_COMMA),
    ("시가대비기관비중(%)", "시가대비_기관매매비중_1주",  18, NUM_FORMAT_COMMA2),
    ("시가대비외국인비중(%)", "시가대비_외국인매매비중_1주", 20, NUM_FORMAT_COMMA2),
    ("섹터",            "섹터",              16,  "@"),
    ("연간매출액(억)",   "연간_매출액",         16,  NUM_FORMAT_COMMA),
    ("연간영업이익(억)", "연간_영업이익",        16,  NUM_FORMAT_COMMA),
    ("영업이익률(%)",    "영업이익률",          14,  NUM_FORMAT_COMMA2),
    ("연간ROE(%)",       "연간_ROE",            12,  NUM_FORMAT_COMMA2),
    ("부채비율(%)",      "부채비율",            12,  NUM_FORMAT_COMMA2),
    ("유동비율(%)",      "유동비율",            12,  NUM_FORMAT_COMMA2),
    ("매출액증가율(%)",  "매출액_증가율",        16,  NUM_FORMAT_COMMA2),
    ("영업이익증가율(%)", "영업이익_증가율",     18,  NUM_FORMAT_COMMA2),
]


def _write_stock_sheet(ws, df: pd.DataFrame, sheet_title: str):
    """종목 데이터 시트 작성"""
    # 타이틀 행
    ws.merge_cells(f"A1:{get_column_letter(len(STOCK_COLUMNS))}1")
    ws["A1"] = sheet_title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # 헤더 행
    for col_idx, (header, src_col, width, fmt) in enumerate(STOCK_COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 18

    # 데이터 행
    for row_idx, (_, row) in enumerate(df.iterrows(), 3):
        for col_idx, (header, src_col, width, fmt) in enumerate(STOCK_COLUMNS, 1):
            value = row.get(src_col) if src_col in df.columns else None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if fmt != "@" and value is not None:
                cell.number_format = fmt

    # 틀 고정
    ws.freeze_panes = "A3"


def _write_market_sheet(ws, processed: dict):
    """증시정보 탭 작성"""
    ws["A1"] = "증시 정보"
    ws["A1"].font = Font(bold=True, size=13)

    headers = ["구분", "종가(p)", "주간등락(p)", "주간등락률(%)", "일간등락(p)", "일간등락률(%)",
               "외국인 주간순매수(억)", "기관 주간순매수(억)", "개인 주간순매수(억)"]

    market_info = processed.get("market_info", {})
    investor_ranks = processed.get("investor_ranks", {})

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    row_num = 3
    for market_label in ["KOSPI", "KOSDAQ"]:
        # market_info["KOSPI_index"] 는 dict {"종가": ..., "전일대비": ..., "등락률": ...}
        idx = market_info.get(f"{market_label}_index", {})

        index_close = idx.get("종가")
        daily_pt = idx.get("전일대비")
        daily_pct = idx.get("등락률")
        # 주간 등락 = DB 이력에서 계산된 값 (없으면 None)
        weekly_pt = idx.get("주간등락(p)")
        weekly_pct = idx.get("주간등락률(%)")

        # 투자자별 순매수
        # 1순위: KIS investor_ranks (현재 API 빈 응답으로 항상 없음)
        fore_net = inst_net = indiv_net = None
        inv_df = investor_ranks.get(market_label, pd.DataFrame())
        if not inv_df.empty:
            if "외국인순매수금액" in inv_df.columns:
                fore_net = float(inv_df["외국인순매수금액"].sum())
            if "기관순매수금액" in inv_df.columns:
                inst_net = float(inv_df["기관순매수금액"].sum())

        # 2순위 폴백: 섹터별 집계 합산 (상위 200종목 기준 근사값)
        if fore_net is None or inst_net is None:
            sec_df = processed.get(
                f"{'kospi' if market_label == 'KOSPI' else 'kosdaq'}_sector",
                pd.DataFrame(),
            )
            if not sec_df.empty:
                if fore_net is None and "1주외국인매매합(억)" in sec_df.columns:
                    fore_net = float(sec_df["1주외국인매매합(억)"].sum())
                if inst_net is None and "1주기관매매합(억)" in sec_df.columns:
                    inst_net = float(sec_df["1주기관매매합(억)"].sum())

        values = [market_label, index_close, weekly_pt, weekly_pct, daily_pt, daily_pct,
                  fore_net, inst_net, indiv_net]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx > 1 and val is not None:
                cell.number_format = NUM_FORMAT_COMMA2

        row_num += 1

    # 컬럼 너비
    widths = [10, 12, 14, 14, 12, 14, 20, 20, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_etf_sheet(ws, etf_df, title_suffix: str = ""):
    """ETF 수급 탭. 4주 추이는 markdown 섹션과 같은 표기로 문자열 한 칸에 담는다."""
    from modules.etf_section import format_trend

    title = f"ETF 수급 ({title_suffix})" if title_suffix else "ETF 수급"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")

    if etf_df is None or not isinstance(etf_df, pd.DataFrame) or etf_df.empty:
        ws["A2"] = "ETF 수급 데이터 없음"
        return

    has_trend = "기관4주추이" in etf_df.columns and "외국인4주추이" in etf_df.columns
    headers = ["티커", "종목명", "기관(억)", "외국인(억)"]
    if has_trend:
        headers += ["기관 4주 추이", "외국인 4주 추이"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # 기관 순매수 큰 순 — markdown 섹션의 '기관 순매수 상위'와 같은 기준으로 읽히게 한다.
    ordered = etf_df.sort_values("1주기관매매", ascending=False, na_position="last")
    for row_idx, (_, row) in enumerate(ordered.iterrows(), 3):
        values = [
            row.get("티커"),
            row.get("종목명"),
            row.get("1주기관매매"),
            row.get("1주외국인매매"),
        ]
        if has_trend:
            values += [format_trend(row.get("기관4주추이")), format_trend(row.get("외국인4주추이"))]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx in (3, 4) and value is not None:
                cell.number_format = NUM_FORMAT_COMMA2

    for column, width in zip("ABCDEF", (10, 26, 14, 14, 30, 30)):
        ws.column_dimensions[column].width = width


def _write_cap_weight_sheet(ws, cap_df: pd.DataFrame, market: str):
    """시가 구간별 비중 탭 작성"""
    title = f"{market} 시가총액 구간별 비중"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["구간", "종목수", "구간시총(억)", "전체시총(억)", "비중(%)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for row_idx, (_, row) in enumerate(cap_df.iterrows(), 3):
        vals = [
            row.get("구간"),
            row.get("종목수"),
            row.get("구간시총(억)"),
            row.get("전체시총(억)"),
            row.get("비중(%)"),
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx >= 3 and val is not None:
                cell.number_format = NUM_FORMAT_COMMA2

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12


# ─────────────────────────────────────────
# 메인 저장 함수
# ─────────────────────────────────────────

def save_excel(processed: dict, config: dict) -> str:
    """Excel 파일 저장 (5탭)"""
    import openpyxl
    from openpyxl import Workbook

    base_date = processed.get("base_date", "")
    prefix = config.get("output", {}).get("excel_prefix", "주가자금동향")
    excel_dir = Path(config.get("output", {}).get("excel_dir", "./data"))
    excel_dir.mkdir(parents=True, exist_ok=True)

    filename = f"{prefix}_{base_date}.xlsx"
    filepath = excel_dir / filename

    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # 시트 타이틀 결정
    period_label = processed.get("period_label", "")
    is_midweek = processed.get("is_midweek", False)
    title_suffix = period_label if is_midweek and period_label else base_date

    # 탭 1: 코스피
    ws_kospi = wb.create_sheet("코스피")
    df_kospi = processed.get("kospi", pd.DataFrame())
    _write_stock_sheet(ws_kospi, df_kospi, f"코스피 전종목 ({title_suffix})")

    # 탭 2: 코스닥
    ws_kosdaq = wb.create_sheet("코스닥")
    df_kosdaq = processed.get("kosdaq", pd.DataFrame())
    _write_stock_sheet(ws_kosdaq, df_kosdaq, f"코스닥 전종목 ({title_suffix})")

    # 탭 3: 증시정보
    ws_market = wb.create_sheet("증시정보")
    _write_market_sheet(ws_market, processed)

    # 탭 4: 코스피 시총비중
    ws_kospi_cap = wb.create_sheet("코스피 시총비중")
    cap_kospi = processed.get("kospi_cap_weight", pd.DataFrame())
    _write_cap_weight_sheet(ws_kospi_cap, cap_kospi, "KOSPI")

    # 탭 5: 코스닥 시총비중
    ws_kosdaq_cap = wb.create_sheet("코스닥 시총비중")
    cap_kosdaq = processed.get("kosdaq_cap_weight", pd.DataFrame())
    _write_cap_weight_sheet(ws_kosdaq_cap, cap_kosdaq, "KOSDAQ")

    # 탭 6: ETF 수급
    ws_etf = wb.create_sheet("ETF")
    _write_etf_sheet(ws_etf, processed.get("etf_flows"), title_suffix)

    wb.save(filepath)
    logger.info(f"Excel 저장 완료: {filepath}")
    return str(filepath)
